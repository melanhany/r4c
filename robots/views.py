from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count
from openpyxl import Workbook
from datetime import datetime, timedelta
import json

from helpers.helpers import validate_json_data
from .signals import robot_created
from .forms import RobotForm
from .models import Robot

# Create your views here.


def create_robot_instance(data) -> RobotForm:
    """
    Create an instance of the RobotForm with the given data.

    Args:
        data (dict): A dictionary containing data for creating an order instance.

    Returns:
        RobotForm: An instance of the RobotForm.
    """

    robot_data = {
        "serial": "{}-{}".format(data["model"], data["version"]),
        "model": data["model"],
        "version": data["version"],
        "created": data["created"],
    }
    return RobotForm(robot_data)


@method_decorator(csrf_exempt, name="dispatch")
def post_robot(request) -> JsonResponse:
    """
    View function for creating a new order via HTTP POST request.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        JsonResponse: A JSON response containing the result of the operation.
    """
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode("utf-8"))

            required_fields = {"model", "version", "created"}
            missing_fields = validate_json_data(data, required_fields)

            if missing_fields:
                missing_fields_list = ", ".join(missing_fields)
                response = {
                    "message": f"Missing required field(s) in JSON data: {missing_fields_list}"
                }
                return JsonResponse(response, status=400)

            form = create_robot_instance(data)
            if form.is_valid():
                robot = form.save()

                robot_created.send(Robot, robot=robot)
                response = {"message": f"New robot added with id: {robot.id}"}
                return JsonResponse(response, status=201)
            else:
                errors = form.errors
                response = {"message": "Validation error", "errors": errors}
                return JsonResponse(response, status=400)

        except json.JSONDecodeError:
            response = {"message": "Invalid JSON format in the request body."}
            return JsonResponse(response, status=400)
    else:
        response = {"message": "Invalid request."}
        return JsonResponse(response, status=400)


@method_decorator(csrf_exempt, name="dispatch")
def generate_weekly_report(request):
    """
    View function for generating a weekly report in Excel format.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: An HTTP response containing the generated Excel report.

    This view function performs the following actions:
    1. Creates a new Excel workbook using the `openpyxl` library.
    2. Determines the current date as `end_date` and calculates `start_date` as 7 days before `end_date`.
    3. Retrieves a list of unique robot models from the database.
    4. Iterates over each unique model to create a separate worksheet for each model in the workbook.
    5. Adds headers for "Модель" (Model), "Версия" (Version), and "Количество за неделю" (Count for the Week) to each worksheet.
    6. Queries the database to retrieve data about the count of each robot model/version created in the past week.
    7. Populates each worksheet with the retrieved data.
    8. Removes the default "Sheet" from the workbook.
    9. Prepares an HTTP response with the appropriate content type for an Excel file.
    10. Sets the content disposition header to specify the filename for the generated Excel report.
    11. Saves the workbook to the response content.
    12. Returns the HTTP response with the generated Excel report as an attachment.

    Example Usage:
    This view is typically accessed via an HTTP request, such as a GET request. When accessed, it generates a weekly
    report that provides insights into the number of robot models and versions created during the past week.
    """

    wb = Workbook()

    end_date = datetime.now()  # Current datetime
    start_date = end_date - timedelta(days=7)  # 7 days before current datetime
    unique_models = Robot.objects.values_list("model", flat=True).distinct()
    for model in unique_models:
        ws = wb.create_sheet(title=model)

        ws["A1"] = "Модель"
        ws["B1"] = "Версия"
        ws["C1"] = "Количество за неделю"

        model_version_data = (
            Robot.objects.filter(model=model)
            .filter(created__range=(start_date, end_date))
            .values("model", "version")
            .annotate(count_for_week=Count("version"))
            .order_by("model", "version")
        )

        for i, data in enumerate(
            model_version_data, start=2
        ):  # Start from row 2, assuming headers in row 1
            ws.cell(row=i, column=1, value=data["model"])
            ws.cell(row=i, column=2, value=data["version"])
            ws.cell(row=i, column=3, value=data["count_for_week"])
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = f"attachment; filename=report {start_date.date()}_{end_date.date()}.xlsx"

    # Save the workbook to the response content
    wb.save(response)

    return response
